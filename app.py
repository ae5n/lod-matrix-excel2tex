import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io
import zipfile
import tempfile
import os
import re

def sanitize_filename(name):
    sanitized = re.sub(r'[<>:"/\\|?*]', '_', str(name))
    sanitized = re.sub(r'\s+', '_', sanitized.strip())
    return sanitized

def escape_latex(text):
    if not text:
        return ""

    text_str = str(text)

    def _escape_chars(s):
        return (
            s.replace("\\", "\\textbackslash{}")
            .replace("&", "\\&")
            .replace("%", "\\%")
            .replace("$", "\\$")
            .replace("#", "\\#")
            .replace("_", "\\_")
            .replace("{", "\\{")
            .replace("}", "\\}")
            .replace("~", "\\textasciitilde{}")
            .replace("^", "\\textasciicircum{}")
        )

    if "•" in text_str:
        text_str = text_str.replace("\n", " ")
        items = [item.strip() for item in text_str.split("•") if item.strip()]
        escaped_items = ["• " + _escape_chars(item) for item in items]
        return " \\newline ".join(escaped_items)
    else:
        return _escape_chars(text_str)

def col_letter_to_index(letter):
    return ord(letter.upper()) - ord("A")

def process_worksheet(ws, sheet_name):
    excluded_indices = set(col_letter_to_index(c) for c in EXCLUDED_COLUMNS)

    header_title = escape_latex(ws["A1"].value)
    
    all_letters = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
    final_letters = [l for l in all_letters if l not in EXCLUDED_COLUMNS]
    final_indices = [col_letter_to_index(l) for l in final_letters]

    headers = []
    for letter in final_letters:
        cell_ref = f"{letter}2"
        header_value = ws[cell_ref].value
        headers.append(escape_latex(header_value))

    tabular_parts = []
    for l in final_letters:
        if l in {"A", "B"}:
            tabular_parts.append(COLUMN_WIDTHS[l])
        else:
            tabular_parts.append(COLUMN_WIDTHS["C"])
    tabular_spec = "|" + "|".join(tabular_parts) + "|"

    color_definition = "% Add these packages to your LaTeX document preamble:\n% \\usepackage{array}\n% \\usepackage{xcolor}\n% \\usepackage{colortbl}\n% \\usepackage{longtable}\n\\definecolor{headercolor}{HTML}{00ACD2}\n"

    # Simplified header structure - no multirow
    header_latex = f"""
\\scriptsize
\\begin{{longtable}}{{{tabular_spec}}}
\\hline
\\rowcolor{{headercolor}}\\multicolumn{{{len(final_letters)}}}{{|c|}}{{\\parbox[c][4ex][c]{{\\linewidth}}{{\\centering\\textcolor{{white}}{{\\textbf{{\\normalsize{{{header_title}}}}}}}}}}} \\\\
\\hline
\\rowcolor{{headercolor}}{" & ".join([f"\\textcolor{{white}}{{\\textbf{{{header}}}}}" for header in headers])} \\\\
\\hline
\\endfirsthead

\\hline
\\rowcolor{{headercolor}}\\multicolumn{{{len(final_letters)}}}{{|c|}}{{\\parbox[c][4ex][c]{{\\linewidth}}{{\\centering\\textcolor{{white}}{{\\textbf{{\\normalsize{{{header_title}}} (continued)}}}}}}}} \\\\
\\hline
\\rowcolor{{headercolor}}{" & ".join([f"\\textcolor{{white}}{{\\textbf{{{header}}}}}" for header in headers])} \\\\
\\hline
\\endhead

\\hline
\\multicolumn{{{len(final_letters)}}}{{|r|}}{{\\textit{{Continued on next page...}}}} \\\\
\\hline
\\endfoot

\\hline
\\endlastfoot
"""

    body_rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        original_filtered = [row[i] if i < len(row) else None for i in final_indices]
        row_has_content = any(
            cell is not None and str(cell).strip() 
            for cell in original_filtered
        )
        
        if not row_has_content:
            continue
            
        escaped = [escape_latex(cell) for cell in row]
        filtered = [escaped[i] for i in final_indices]
        
        has_content_only_in_A = False
        if "A" not in EXCLUDED_COLUMNS:
            first_col_has_content = original_filtered[0] is not None and str(original_filtered[0]).strip()
            other_cols_empty = all(
                not (cell is not None and str(cell).strip()) 
                for cell in original_filtered[1:]
            )
            has_content_only_in_A = first_col_has_content and other_cols_empty
        
        if has_content_only_in_A:
            body_rows.append("\\rowcolor{headercolor}" + " & ".join(filtered) + " \\\\")
        else:
            body_rows.append(" & ".join(filtered) + " \\\\")

    body_latex = "\n\\hline\n".join(body_rows)
    latex_table = header_latex + body_latex + "\n\\end{longtable}"
    
    return color_definition + "\n" + latex_table

def create_zip_file(latex_files):
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for filename, content in latex_files.items():
            zip_file.writestr(filename, content)
    zip_buffer.seek(0)
    return zip_buffer

def main():
    st.set_page_config(
        page_title="LOD Matrix Excel2TeX",
        page_icon="📊",
        layout="wide"
    )
    
    st.title("LOD Matrix Excel2TeX")
    st.markdown("Upload an Excel file to generate LaTeX tables for each worksheet")
    
    with st.sidebar:
        st.header("⚙️ Configuration")
        
        st.subheader("Excluded Columns")
        excluded_columns_input = st.multiselect(
            "Select columns to exclude (default: B, C, D, E)",
            options=["A", "B", "C", "D", "E", "F", "G", "H", "I"],
            default=["B", "C", "D", "E"]
        )
        
        st.subheader("Column Widths")
        col_width_A = st.number_input(
            "Column A width", 
            min_value=1.0, 
            max_value=10.0, 
            value=4.0, 
            step=0.5
        )
        col_width_B = st.number_input(
            "Column B width", 
            min_value=1.0, 
            max_value=10.0, 
            value=4.0, 
            step=0.5,
        )
        col_width_C = st.number_input(
            "Columns C-I width", 
            min_value=1.0, 
            max_value=10.0, 
            value=2.0, 
            step=0.5
        )
        
        global EXCLUDED_COLUMNS, COLUMN_WIDTHS
        EXCLUDED_COLUMNS = set(excluded_columns_input)
        COLUMN_WIDTHS = {
            "A": f"m{{{col_width_A}cm}}",
            "B": f"m{{{col_width_B}cm}}",
            "C": f">{{\\centering\\arraybackslash}}m{{{col_width_C}cm}}",
        }
    
    uploaded_file = st.file_uploader(
        "Choose an Excel file", 
        type=['xlsx', 'xls'],
        help="Upload your Excel file containing the BIMx data"
    )
    
    if uploaded_file is not None:
        try:
            wb = load_workbook(uploaded_file)
            st.success(f"✅ File loaded successfully! Found {len(wb.sheetnames)} worksheets.")
            
            with st.expander("📋 Worksheets found in your file"):
                for i, sheet_name in enumerate(wb.sheetnames, 1):
                    st.write(f"{i}. {sheet_name}")
            
            if st.button("🚀 Generate LaTeX Tables", type="primary"):
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                latex_files = {}
                
                for i, sheet_name in enumerate(wb.sheetnames):
                    status_text.text(f"Processing worksheet: {sheet_name}")
                    
                    try:
                        ws = wb[sheet_name]
                        latex_content = process_worksheet(ws, sheet_name)
                        filename = f"{sanitize_filename(sheet_name)}.tex"
                        latex_files[filename] = latex_content
                        
                        progress_bar.progress((i + 1) / len(wb.sheetnames))
                        
                    except Exception as e:
                        st.error(f"Error processing worksheet '{sheet_name}': {str(e)}")
                        continue
                
                status_text.text("✅ Processing complete!")
                
                if latex_files:
                    st.success(f"🎉 Generated {len(latex_files)} LaTeX files!")
                    
                    col1, col2 = st.columns([2, 1])
                    
                    with col1:
                        st.subheader("📁 Generated Files")
                        
                        for filename in latex_files.keys():
                            with st.expander(f"📄 {filename} - Click to view full LaTeX code"):
                                st.code(latex_files[filename], language="latex")
                    
                    with col2:
                        st.subheader("💾 Download")
                        
                        st.write("**Individual Files:**")
                        for filename, content in latex_files.items():
                            st.download_button(
                                label=f"📄 {filename}",
                                data=content,
                                file_name=filename,
                                mime="text/plain",
                                key=f"download_{filename}"
                            )
                        
                        st.write("---")
                        
                        st.write("**All Files:**")
                        zip_file = create_zip_file(latex_files)
                        st.download_button(
                            label="📦 Download All as ZIP",
                            data=zip_file,
                            file_name="latex_tables.zip",
                            mime="application/zip"
                        )
                        
                        with st.expander("📖 LaTeX Usage Instructions"):
                            st.markdown("""
                            To use these LaTeX tables:
                            
                            1. **Add required packages** to your LaTeX document preamble:
                            ```latex
                            \\usepackage{array}
                            \\usepackage{xcolor}
                            \\usepackage{colortbl}
                            \\usepackage{longtable}
                            ```
                            
                            2. **Include the .tex file** in your document:
                            ```latex
                            \\input{filename.tex}
                            ```
                            
                            3. **Compile** with pdflatex or xelatex
                            """)
                else:
                    st.error("No LaTeX files were generated. Please check your Excel file format.")
        
        except Exception as e:
            st.error(f"Error loading file: {str(e)}")
            st.info("Please make sure you uploaded a valid Excel file (.xlsx or .xls)")
    
    else:
        pass

if __name__ == "__main__":
    main() 